"""Excel (.xlsx) reader."""

from __future__ import annotations

import sys
from pathlib import Path

from officecat.models import Sheet, Workbook

DEFAULT_ROW_CAP = 500


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _col_letter(index: int) -> str:
    """Convert 0-based column index to Excel-style letter (A, B, ... Z, AA, ...)."""
    result = ""
    i = index
    while True:
        result = chr(65 + i % 26) + result
        i = i // 26 - 1
        if i < 0:
            break
    return result


def read_xlsx(
    path: Path,
    *,
    head: int | None = None,
    sheet: str | None = None,
    headers: int = 1,
    show_all: bool = False,
) -> Workbook:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except InvalidFileException:
        print(
            f"Error: '{path.name}' appears to be corrupt or is not a valid xlsx file.",
            file=sys.stderr,
        )
        sys.exit(3)
    except Exception as e:
        msg = str(e).lower()
        if "password" in msg or "encrypt" in msg:
            print(
                f"Error: '{path.name}' is password-protected. "
                f"officecat cannot open encrypted files.",
                file=sys.stderr,
            )
            sys.exit(3)
        raise

    try:
        all_sheet_names = wb.sheetnames

        if sheet is not None:
            # Try as 1-based index
            try:
                idx = int(sheet)
                if 1 <= idx <= len(all_sheet_names):
                    sheets_to_read = [wb[all_sheet_names[idx - 1]]]
                else:
                    print(
                        f"Error: Sheet index {idx} out of range. "
                        f"Available sheets: {', '.join(all_sheet_names)}",
                        file=sys.stderr,
                    )
                    sys.exit(1)
            except ValueError:
                if sheet in all_sheet_names:
                    sheets_to_read = [wb[sheet]]
                else:
                    print(
                        f"Error: Sheet '{sheet}' not found. "
                        f"Available sheets: {', '.join(all_sheet_names)}",
                        file=sys.stderr,
                    )
                    sys.exit(1)
        else:
            sheets_to_read = [wb[name] for name in all_sheet_names]

        row_cap = head if head is not None else (None if show_all else DEFAULT_ROW_CAP)

        result_sheets = []
        for ws in sheets_to_read:
            total_rows = ws.max_row or 0
            col_count = ws.min_column  # fallback
            all_rows: list[list[str]] = []
            header_row: list[str] = []

            for i, row in enumerate(ws.iter_rows(), start=1):
                if headers > 0 and i == headers:
                    header_row = [_format_cell(cell.value) for cell in row]
                    col_count = len(header_row)
                    continue

                values = [_format_cell(cell.value) for cell in row]
                all_rows.append(values)

                if row_cap is not None and len(all_rows) >= row_cap:
                    break

            if not header_row:
                # Use Excel-style column letters
                if all_rows:
                    col_count = max(len(r) for r in all_rows)
                else:
                    col_count = ws.max_column or 0
                header_row = [_col_letter(i) for i in range(col_count)]

            # Adjust total_rows to exclude the header row
            effective_total = max(0, total_rows - (1 if headers > 0 else 0))

            result_sheets.append(
                Sheet(
                    name=ws.title,
                    headers=header_row,
                    rows=all_rows,
                    total_rows=effective_total,
                )
            )

        return Workbook(
            source=path.name,
            sheets=result_sheets,
            sheet_names=all_sheet_names,
        )
    finally:
        wb.close()
